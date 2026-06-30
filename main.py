import os
import sys
import json
import asyncio

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import time
import base64
from datetime import datetime, timedelta
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from collections import defaultdict
from playwright.async_api import async_playwright
import re
import io
import openpyxl
import hashlib
import logging
from logging.handlers import RotatingFileHandler

# === LOGGING SETUP ===
LOG_DIR = os.path.dirname(os.path.abspath(__file__))
log_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, "activity.log"),
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3,
    encoding="utf-8"
)
log_handler.setFormatter(logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
))
logger = logging.getLogger("gplx")
logger.setLevel(logging.INFO)
logger.addHandler(log_handler)

import database as db
from app_ocr import doc_captcha_ddddocr, ocr  # Giữ lại hàm OCR cũ

app = FastAPI(title="GPLX Verification System")

# Cấu hình tĩnh
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(BASE_DIR, "static")
os.makedirs(static_dir, exist_ok=True)
app.mount("/static", StaticFiles(directory=static_dir), name="static")

# --- RATE LIMITER ---
RATE_LIMIT_MINUTES = 1
MAX_REQUESTS_PER_MINUTE = 200  # Nới rộng cho tra cứu hàng loạt
ip_requests = defaultdict(list)

def check_rate_limit(client_ip: str) -> tuple[bool, str]:
    now = datetime.now()
    # Lọc bỏ các request cũ
    ip_requests[client_ip] = [req_time for req_time in ip_requests[client_ip] if now - req_time < timedelta(minutes=RATE_LIMIT_MINUTES)]
    
    if len(ip_requests[client_ip]) >= MAX_REQUESTS_PER_MINUTE:
        return False, f"Bạn đã vượt quá giới hạn {MAX_REQUESTS_PER_MINUTE} yêu cầu/phút. Vui lòng thử lại sau."
    
    ip_requests[client_ip].append(now)
    return True, ""

# --- API ENDPOINTS ---

@app.get("/")
async def get_index():
    return FileResponse(os.path.join(static_dir, "index.html"))

# Mật khẩu quản trị (bạn có thể thay đổi tại đây)
ADMIN_SECRET = "gplx2025"

@app.get("/api/export-excel")
async def export_excel(key: str = ""):
    if key != ADMIN_SECRET:
        logger.warning(f"Export bị từ chối - sai mật khẩu")
        return JSONResponse({"error": "Không có quyền truy cập."}, status_code=403)
    logger.info(f"Admin xuất file Excel")
    history = db.get_all_history()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lich Su GPLX"
    
    # Header gọn gàng
    ws.append(["Họ Tên", "Số GPLX", "Loại Bằng", "Ngày Cấp", "Thời Hạn", "Trạng Thái"])
    
    for row in history:
        name = row.get("name", "")
        gplx = row["gplx"]
        loai_bang = ""
        ngay_cap = ""
        thoi_han = ""
        status_map = {"success": "Thành công", "not_found": "Không tìm thấy", "error": "Lỗi"}
        trang_thai = status_map.get(row["status"], row["status"])
        
        # Trích xuất thông tin từ data JSON
        verified_data = db.get_verified_data(row["gplx"])
        if verified_data:
            if isinstance(verified_data, dict):
                for k, v in verified_data.items():
                    kl = k.lower()
                    val = str(v).strip() if v else ""
                    if not val:
                        continue
                    # Tên
                    if not name and ("tên" in kl or "name" in kl):
                        name = val
                    # Loại bằng / Hạng
                    if "hạng" in kl or "loại" in kl or "class" in kl or "hang" in kl:
                        loai_bang = val
                    # Ngày cấp
                    if "ngày cấp" in kl or "cấp ngày" in kl or "issue" in kl or "ngaycap" in kl:
                        ngay_cap = val
                    # Thời hạn / Có giá trị đến
                    if "thời hạn" in kl or "hạn" in kl or "giá trị" in kl or "expir" in kl or "valid" in kl:
                        thoi_han = val
            elif isinstance(verified_data, list):
                for item in verified_data:
                    if isinstance(item, dict):
                        for k, v in item.items():
                            kl = k.lower()
                            val = str(v).strip() if v else ""
                            if not val:
                                continue
                            if not name and ("tên" in kl or "name" in kl):
                                name = val
                            if "hạng" in kl or "loại" in kl or "class" in kl or "hang" in kl:
                                loai_bang = val
                            if "ngày cấp" in kl or "cấp ngày" in kl or "issue" in kl or "ngaycap" in kl:
                                ngay_cap = val
                            if "thời hạn" in kl or "hạn" in kl or "giá trị" in kl or "expir" in kl or "valid" in kl:
                                thoi_han = val
        
        ws.append([name, gplx, loai_bang, ngay_cap, thoi_han, trang_thai])
    
    # Save to temp file
    export_path = os.path.join(BASE_DIR, "Lich_su_GPLX.xlsx")
    wb.save(export_path)
    
    return FileResponse(
        export_path, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename="Lich_su_GPLX.xlsx"
    )

@app.post("/api/upload-gplx")
async def upload_gplx_image(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        
        # Gọi ddddocr để quét ảnh
        text = ocr.classification(contents)
        
        # Dùng Regex để tìm Số GPLX và Ngày sinh
        # Số GPLX PET thường là 12 chữ số
        gplx_match = re.search(r'\b\d{12}\b', text)
        gplx_number = gplx_match.group(0) if gplx_match else ""
        
        # Ngày sinh định dạng dd/mm/yyyy
        dob_match = re.search(r'\b\d{2}[/.-]\d{2}[/.-]\d{4}\b', text)
        dob = dob_match.group(0).replace('-', '/').replace('.', '/') if dob_match else ""
        
        return JSONResponse({"status": "success", "gplx": gplx_number, "dob": dob, "raw_text": text})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)})

@app.post("/api/upload-bulk")
async def upload_bulk(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        records = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue # Skip header
            if not row[0] or not row[1]:
                continue
                
            gplx = str(row[0]).strip()
            dob = str(row[1]).strip()
            loai = str(row[2]).strip().upper() if len(row) > 2 and row[2] else "PET"
            
            records.append({
                "gplx": gplx,
                "dob": dob,
                "loai_bang": "PET" if "PET" in loai or "2" in loai else "OLD"
            })
            
        return JSONResponse({"status": "success", "data": records})
    except Exception as e:
        return JSONResponse({"status": "error", "message": f"Không thể đọc file: {str(e)}"})

# --- WEBSOCKET FLOW ---

async def parse_response_to_data(raw_response: str):
    """Phân tích raw_response (JSON hoặc text) thành dict {status, message, data, is_captcha_error}"""
    if raw_response.strip() == "BotDetect":
        return {"status": "captcha_error", "message": "Sai mã captcha (BotDetect)", "is_captcha_error": True}
        
    try:
        data = json.loads(raw_response)
        if data:
            if isinstance(data, dict):
                if 'data' in data and data['data']:
                    return {"status": "success", "data": data['data'], "is_captcha_error": False}
                
                msg = data.get('message', '').lower()
                if 'không tìm thấy' in msg or 'not found' in msg:
                    return {"status": "not_found", "message": msg, "is_captcha_error": False}
                
                if 'captcha' in msg or 'mã bảo mật' in msg:
                    return {"status": "captcha_error", "message": "Sai Captcha", "is_captcha_error": True}
                
                if data:
                    return {"status": "success", "data": data, "is_captcha_error": False}
            elif isinstance(data, list):
                if data:
                    return {"status": "success", "data": data, "is_captcha_error": False}
                else:
                    return {"status": "not_found", "message": "Không tìm thấy dữ liệu", "is_captcha_error": False}
            else:
                return {"status": "success", "data": data, "is_captcha_error": False}
        return {"status": "not_found", "message": "Không tìm thấy dữ liệu", "is_captcha_error": False}
    except json.JSONDecodeError:
        if 'không tìm thấy' in raw_response.lower():
            return {"status": "not_found", "message": "Không tìm thấy thông tin", "is_captcha_error": False}
        elif 'thành công' in raw_response.lower():
            return {"status": "success", "data": raw_response, "is_captcha_error": False}
        elif len(raw_response) > 10:
            return {"status": "success", "data": raw_response, "is_captcha_error": False}
        return {"status": "unknown", "message": "Lỗi không xác định: " + raw_response[:50], "is_captcha_error": False}

@app.websocket("/ws/verify")
async def verify_gplx_ws(websocket: WebSocket):
    await websocket.accept()
    
    # 1. Nhận thông tin ban đầu
    try:
        init_data = await websocket.receive_json()
    except WebSocketDisconnect:
        return
        
    gplx = init_data.get("gplx", "").replace(" ", "")
    dob = init_data.get("dob", "").replace(" ", "")
    loai_bang = init_data.get("loai_bang", "PET")
    client_ip = websocket.client.host
    
    logger.info(f"[{client_ip}] Tra cứu: GPLX={gplx}, DOB={dob}, Loại={loai_bang}")
    
    # Rate Limit
    is_allowed, rl_msg = check_rate_limit(client_ip)
    if not is_allowed:
        await websocket.send_json({"type": "error", "message": rl_msg})
        await websocket.close()
        return

    # Check Database — phải khớp cả GPLX VÀ ngày sinh
    if db.is_gplx_verified(gplx):
        stored_dob = db.get_stored_dob(gplx)
        if stored_dob and stored_dob == dob:
            verified_data = db.get_verified_data(gplx)
            logger.info(f"[{client_ip}] Kết quả từ DB: GPLX={gplx} - Thành công")
            await websocket.send_json({
                "type": "success", 
                "source": "database",
                "message": "Giấy phép này đã được xác thực trước đó.",
                "data": verified_data
            })
            await websocket.close()
            return
        elif stored_dob and stored_dob != dob:
            logger.warning(f"[{client_ip}] Sai ngày sinh cho GPLX={gplx}")
            await websocket.send_json({
                "type": "error",
                "source": "database", 
                "message": "Ngày sinh không khớp với thông tin đã xác thực trước đó."
            })
            await websocket.close()
            return

    await websocket.send_json({"type": "status", "message": "Đang khởi tạo trình duyệt ẩn..."})
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage", # Chống crash do thiếu bộ nhớ /dev/shm trên Docker
                "--disable-gpu",
                "--single-process" # Tối ưu RAM cho gói Free của Render
            ]
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            locale="vi-VN"
        )
        page = await context.new_page()
        
        try:
            await page.goto("https://gplx.csgt.bocongan.gov.vn/", wait_until="networkidle")
            security_token = await page.locator("input[name='securityToken']").get_attribute("value")
            # 1: Giấy Cũ, 2: PET (Có thời hạn), 3: PET (Không thời hạn)
            # Theo code cũ: PET = 2, Cũ = 1
            choose_gplx = "1" if loai_bang == "OLD" else "2"
            
            async def get_and_read_captcha(save_path="captcha_real.png"):
                captcha_selector = ".img-cap-mobile a.captcha-refresh img"
                await page.wait_for_selector(captcha_selector, timeout=5000)
                captcha_element = await page.query_selector(captcha_selector)
                await captcha_element.screenshot(path=save_path)
                return doc_captcha_ddddocr(save_path)
            
            async def submit_form(cap_code: str):
                payload_data = {
                    "type": "",
                    "fields[formTypeId]": "565f96637f8b9af6558b4567",
                    "fields[chooseGPLX]": str(choose_gplx),
                    "fields[codeGPLX]": str(gplx),
                    "fields[birthDate]": str(dob),
                    "fields[birthDateType2]": "",
                    "captcha_code": str(cap_code).lower().strip(),
                    "securityToken": str(security_token),
                    "submitFormId": "8",
                    "moduleId": "8"
                }
                
                api_url = "/api/Project/GPLX/ApiSearchGPLX/sendRequest?site=2005782"
                raw_response = await page.evaluate(
                    f"""async (formDataObj) => {{
                        const searchParams = new URLSearchParams();
                        for (const key in formDataObj) {{
                            searchParams.append(key, formDataObj[key]);
                        }}
                        const res = await fetch('{api_url}', {{
                            method: 'POST',
                            headers: {{
                                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                                'Accept': 'application/json, text/plain, */*'
                            }},
                            body: searchParams.toString()
                        }});
                        return res.text();
                    }}""",
                    payload_data
                )
                return await parse_response_to_data(raw_response)

            # === AUTOMATIC OCR LOOP (MAX 2 TIMES) ===
            auto_attempts = 0
            max_auto = 2
            success_result = None
            
            while auto_attempts < max_auto:
                auto_attempts += 1
                await websocket.send_json({"type": "status", "message": f"Đang tự động đọc Captcha (Lần {auto_attempts}/{max_auto})..."})
                
                cap_code = await get_and_read_captcha()
                
                if not cap_code:
                    await websocket.send_json({"type": "status", "message": "OCR không đọc được, tải captcha mới..."})
                    continue
                    
                await websocket.send_json({"type": "status", "message": f"Thử Captcha: {cap_code}..."})
                result = await submit_form(cap_code)
                
                if result.get("is_captcha_error"):
                    await websocket.send_json({"type": "status", "message": "Sai Captcha tự động."})
                    await asyncio.sleep(1)
                    continue
                else:
                    success_result = result
                    break
            
            # === MANUAL CAPTCHA IF AUTO FAILED ===
            if not success_result:
                await websocket.send_json({"type": "status", "message": "Giải Captcha tự động thất bại. Yêu cầu nhập tay."})
                
                # Vòng lặp cho phép nhập tay sai rồi nhập lại
                while not success_result:
                    # Lấy captcha mới nhất
                    captcha_selector = ".img-cap-mobile a.captcha-refresh img"
                    await page.wait_for_selector(captcha_selector, timeout=5000)
                    captcha_element = await page.query_selector(captcha_selector)
                    await captcha_element.screenshot(path="captcha_manual.png")
                    
                    with open("captcha_manual.png", "rb") as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                    
                    # Gửi yêu cầu nhập tay
                    await websocket.send_json({
                        "type": "require_manual_captcha",
                        "image_base64": encoded_string,
                        "message": "Vui lòng nhập mã Captcha bên dưới."
                    })
                    
                    try:
                        # Đợi user gửi lên
                        user_resp = await websocket.receive_json()
                        manual_cap = user_resp.get("captcha_code", "")
                        
                        await websocket.send_json({"type": "status", "message": "Đang gửi yêu cầu tra cứu..."})
                        result = await submit_form(manual_cap)
                        
                        if result.get("is_captcha_error"):
                            await websocket.send_json({"type": "status", "message": "Bạn đã nhập sai mã Captcha. Thử lại..."})
                            await asyncio.sleep(0.5)
                            continue
                        else:
                            success_result = result
                            break
                            
                    except WebSocketDisconnect:
                        return # User ngắt kết nối
            
            # === XỬ LÝ KẾT QUẢ ===
            status = success_result.get("status")
            if status == "success":
                db.save_verification(gplx, dob, "success", success_result.get("data"))
                logger.info(f"[{client_ip}] Xác thực thành công: GPLX={gplx}")
                await websocket.send_json({
                    "type": "success",
                    "source": "live",
                    "message": "Tra cứu thành công!",
                    "data": success_result.get("data")
                })
            elif status == "not_found":
                db.save_verification(gplx, dob, "not_found", None)
                logger.info(f"[{client_ip}] Không tìm thấy: GPLX={gplx}")
                await websocket.send_json({
                    "type": "error",
                    "source": "live",
                    "message": "Không tìm thấy thông tin trên hệ thống (Cục CSGT)."
                })
            else:
                db.save_verification(gplx, dob, "error", None)
                logger.error(f"[{client_ip}] Lỗi: GPLX={gplx} - {success_result.get('message', '')}")
                await websocket.send_json({
                    "type": "error",
                    "source": "live",
                    "message": success_result.get("message", "Lỗi không xác định.")
                })

        except Exception as e:
            import traceback
            traceback.print_exc()
            await websocket.send_json({"type": "error", "message": f"Lỗi hệ thống: {str(e)}"})
        finally:
            await browser.close()
            
    # Đảm bảo đóng websocket (nếu chưa đóng do error block trên)
    try:
        await websocket.close()
    except Exception:
        pass

if __name__ == "__main__":
    import uvicorn
    # Vô hiệu hóa reload để tránh lỗi subprocess của Uvicorn ghi đè Event Loop trên Windows
    uvicorn.run(app, host="0.0.0.0", port=8001, loop="asyncio")

